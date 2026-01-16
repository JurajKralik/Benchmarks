#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule


COLUMNS = [
	"distribution",
	"n",
	"language",
	"algo",
	"runs",
	"median_ms",
	"iqr_ms",
	"mean_ms",
	"std_ms",
	"speedup_vs_cpp",
	"rel_variability_iqr_over_median",
]

NUMERIC_COLS_3DP = {
	"median_ms",
	"iqr_ms",
	"mean_ms",
	"std_ms",
	"speedup_vs_cpp",
	"rel_variability_iqr_over_median",
}


def read_summary(summary_path: Path) -> list[dict]:
	"""
	Read summary.csv and normalize data types.
	"""
	rows = []
	with summary_path.open(newline="") as f:
		reader = csv.DictReader(f)
		required = {
			"language",
			"algo",
			"distribution",
			"n",
			"runs",
			"median_ms",
			"iqr_ms",
			"mean_ms",
			"std_ms",
		}
		if not required.issubset(reader.fieldnames or []):
			raise SystemExit(f"{summary_path} is missing required columns")

		for row in reader:
			rows.append({
				"language": row["language"],
				"algo": row["algo"],
				"distribution": row["distribution"],
				"n": int(row["n"]),
				"runs": int(row["runs"]),
				"median_ms": float(row["median_ms"]),
				"iqr_ms": float(row["iqr_ms"]),
				"mean_ms": float(row["mean_ms"]),
				"std_ms": float(row["std_ms"]),
			})
	return rows


def compute_speedup(rows: list[dict], baseline_lang: str) -> None:
	"""
	Compute speedup relative to baseline language (default: C++).
	"""
	baseline = {}
	for r in rows:
		if r["language"] == baseline_lang:
			key = (r["distribution"], r["algo"], r["n"])
			baseline[key] = r["median_ms"]

	for r in rows:
		key = (r["distribution"], r["algo"], r["n"])
		base = baseline.get(key)

		if base is None or r["median_ms"] == 0.0:
			r["speedup_vs_cpp"] = None
		else:
			r["speedup_vs_cpp"] = base / r["median_ms"]

		r["rel_variability_iqr_over_median"] = (
			r["iqr_ms"] / r["median_ms"]
			if r["median_ms"] != 0.0 else None
		)


def autosize_columns(ws) -> None:
	"""
	Automatically adjust column widths based on content.
	"""
	for col in ws.columns:
		max_len = 0
		col_letter = col[0].column_letter
		for cell in col:
			value = "" if cell.value is None else str(cell.value)
			max_len = max(max_len, len(value))
		ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)


def add_excel_table(ws, start_row: int, start_col: int,
					end_row: int, end_col: int, name: str) -> None:
	"""
	Add a formatted Excel table to the worksheet.
	"""
	def col_name(n: int) -> str:
		s = ""
		while n:
			n, r = divmod(n - 1, 26)
			s = chr(65 + r) + s
		return s

	ref = f"{col_name(start_col)}{start_row}:{col_name(end_col)}{end_row}"
	table = Table(displayName=name, ref=ref)
	style = TableStyleInfo(
		name="TableStyleMedium9",
		showFirstColumn=False,
		showLastColumn=False,
		showRowStripes=True,
		showColumnStripes=False,
	)
	table.tableStyleInfo = style
	ws.add_table(table)


def format_sheet(ws) -> None:
	"""
	Apply formatting, number formats, and conditional formatting.
	"""
	header_font = Font(bold=True)
	for cell in ws[1]:
		cell.font = header_font
		cell.alignment = Alignment(horizontal="center", vertical="center")

	ws.freeze_panes = "A2"

	headers = [c.value for c in ws[1]]
	for col_idx, name in enumerate(headers, start=1):
		if name in NUMERIC_COLS_3DP:
			for row_idx in range(2, ws.max_row + 1):
				ws.cell(row=row_idx, column=col_idx).number_format = "0.000"

	if "speedup_vs_cpp" in headers:
		j = headers.index("speedup_vs_cpp") + 1
		rng = f"{ws.cell(2, j).coordinate}:{ws.cell(ws.max_row, j).coordinate}"
		ws.conditional_formatting.add(
			rng,
			ColorScaleRule(
				start_type="num", start_value=0.0,
				mid_type="num", mid_value=1.0,
				end_type="num", end_value=2.0,
			)
		)

	autosize_columns(ws)


def main() -> None:
	parser = argparse.ArgumentParser()
	parser.add_argument("--summary", default="results/summary.csv",
						help="Path to summary.csv")
	parser.add_argument("--out", default="results/summary_by_distribution.xlsx",
						help="Output Excel file")
	parser.add_argument("--baseline", default="cpp",
						help="Baseline language for speedup computation")
	parser.add_argument("--include-all", action="store_true",
						help="Include aggregated 'All' sheet")
	args = parser.parse_args()

	summary_path = Path(args.summary)
	if not summary_path.exists():
		raise SystemExit(f"Missing {summary_path}")

	rows = read_summary(summary_path)
	compute_speedup(rows, args.baseline)

	by_dist = defaultdict(list)
	for r in rows:
		by_dist[r["distribution"]].append(r)

	wb = Workbook()
	wb.remove(wb.active)

	if args.include_all:
		ws = wb.create_sheet("All")
		ws.append(COLUMNS)
		for r in sorted(rows, key=lambda x: (x["distribution"], x["n"], x["language"])):
			ws.append([
				r["distribution"], r["n"], r["language"], r["algo"], r["runs"],
				r["median_ms"], r["iqr_ms"], r["mean_ms"], r["std_ms"],
				r["speedup_vs_cpp"], r["rel_variability_iqr_over_median"],
			])
		format_sheet(ws)
		add_excel_table(ws, 1, 1, ws.max_row, len(COLUMNS), "AllResults")

	for dist, dist_rows in sorted(by_dist.items()):
		ws = wb.create_sheet(dist[:31])
		ws.append(COLUMNS)

		for r in sorted(dist_rows, key=lambda x: (x["n"], x["language"])):
			ws.append([
				r["distribution"], r["n"], r["language"], r["algo"], r["runs"],
				r["median_ms"], r["iqr_ms"], r["mean_ms"], r["std_ms"],
				r["speedup_vs_cpp"], r["rel_variability_iqr_over_median"],
			])

		format_sheet(ws)
		add_excel_table(ws, 1, 1, ws.max_row, len(COLUMNS), f"T_{dist}")

	out_path = Path(args.out)
	out_path.parent.mkdir(parents=True, exist_ok=True)
	wb.save(out_path)

	print(f"Excel file written to {out_path}")


if __name__ == "__main__":
	main()
